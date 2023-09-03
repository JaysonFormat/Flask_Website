import secrets
import os
import io
import smtplib
import requests
from project.utility import remove_unused_images,parse_price
from flask_mail import Message, Mail
from PIL import Image
from flask import render_template, request, url_for, flash, redirect, send_file,session, send_from_directory, current_app, session, jsonify, abort
from project import app, db, bcrypt, mail
from project.forms import (
    RegistrationForm, LoginForm, UpdateAccountForm, AppointmentForm, 
    EditUserForm, InventoryForm, UpdatePasswordForm, UserUpdatePasswordForm, EmployeeForm, 
    UpdateEmployee, RequestResetForm,ResetPasswordForm, TimeInForm, TimeOutForm, AttendanceForm, PayrollForm)
from project.models import User,Book_date, Inventory,AuditTrail, Inventory, Employee, Attendance, Payroll
from flask_login import login_user, current_user, logout_user, login_required
from datetime import datetime, time, timedelta,date
from werkzeug.utils import secure_filename
from sqlalchemy import or_, func, extract
from openpyxl.utils import get_column_letter
from functools import wraps
from flask import abort
from flask_login import current_user


#Any happens bad ni update ko sa latest version 2.3.X from 2.2.2
        #Ni update ko sa 2.2.5
#Updating Flask_WTF from 1.0.1 to updated to fix the AttributeError: module 'flask.json' has no attribute 'JSONEncoder' (Fix for now)
        #Rollback ko sa previous version 1.0.1


def deny_customer_staff(f):     #REMOVE @deny_customer_staff sa mga route if may error encountered
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if the user's role is either "Customer" or "Staff"
        if current_user.role in ["Customer", "Staff"]:
            abort(403)  # Raise a 403 Forbidden error if the user has one of the restricted roles
        return f(*args, **kwargs)

    return decorated_function # Mag add for Admins naman for specific routes


@app.route('/')  # HOME PAGE
def index():
    return render_template('Website.html', title='Nail Lounge and Spa')


@app.route('/login/', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            if user.is_active == False:
                flash("Sorry Account is deactivated please contact support", 'danger')
                return redirect(url_for('login'))

            if user.role in ['Customer']:
                login_user(user, remember=form.remember.data)
                session.permanent = True
                session['user_id'] = user.user_id
                next_page = request.args.get('next')
                success_login = AuditTrail(user_id=user.user_id, event_type='Success Login', event_description='User successfully logged in')
                db.session.add(success_login)
                db.session.commit()
                return redirect(next_page) if next_page else redirect(url_for('index'))
            else:
                flash('Access Denied. Please Check Email and Password', 'danger')
                fail_login = AuditTrail(user_id=user.user_id, event_type='Failed Login', event_description='Access denied')
                db.session.add(fail_login)
                db.session.commit()
        else:
            flash('Login Unsuccessful. Please Check Email and Password', 'danger')
            if user:
                fail_login = AuditTrail(user_id=user.user_id, event_type='Failed Login', event_description='Failed to login')
                db.session.add(fail_login)
                db.session.commit()

    return render_template('login.html', title='Login', form=form)


@app.route('/admin_login/', methods=['GET', 'POST'])
def login_admin():
    # if current_user.is_authenticated:
    #     return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            if user.is_active == False:
                flash("Sorry Account is deactivated please contact support", 'danger')
                return redirect(url_for('login_admin'))

            if user.role in ["Admin", "Super_admin"]: 
                login_user(user, remember=form.remember.data)
                next_page = request.args.get('next')
                session.permanent = True
                session['user_id'] = user.user_id
                return redirect(next_page) if next_page else redirect(url_for('adminpage'))

            else:
                flash("Access Denied, You don't have the necessary permission", 'warning')

        else:
            flash('Login Unsuccessful. Please Check Email and Password', 'danger')

    return render_template('admin_login.html', title='Login', form=form)



@app.route('/signup/', methods=['GET', 'POST'])  # Signup Page
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = RegistrationForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user = User(fname=form.fname.data, lname=form.lname.data, email=form.email.data, password=hashed_password)

        
        success_register = AuditTrail(user_id=user.user_id, event_type='Success Register', event_description='User successfully to register')
        db.session.add(success_register)
        db.session.add(user)
        db.session.commit()

        #read the email body from the file
        with open(os.path.join(os.path.dirname(__file__), 'email_body.txt')) as file:
            message_body = file.read()

        # replace any placeholders in the email body with user-specific information
        email_body = message_body.format(form.fname.data)

        # send the confirmation email to the user
        try:
            msg = Message('Thank you for Registering!', sender='nailloungebusiness94@gmail.com', recipients=[form.email.data])
            msg.body = email_body
            mail.send(msg)
            print('already send email')
        except Exception as e:
            print(f"Error sending email: {str(e)}")
            flash('Your account has been created! You are now able to log in', 'success')
        else:
            flash('Your account has been created! You are now able to log in', 'success')

        return redirect(url_for('login'))

    return render_template('Sign_up.html', title='Sign Up', form=form)


@app.route('/terms_and_condition/')
def terms_and_conditions():
    return render_template('terms_and_conditions.html')


@app.route('/service/')  # SERVICE PAGE
def service():
    return render_template('service.html', title='Service')

@app.route('/FAQ/')
def FAQ():
    return render_template('FAQ.html', title='FAQ')

@app.route('/book/')  # Booking PAGE
def booking():
    form = AppointmentForm()
    return render_template('book.html', title='Booking', form=form)

@app.route('/appointment_time_zapote/', methods=['GET', 'POST']) #Calendar Appointment zapote
def app_time_zap():
    date = request.args.get('date')

    if date:
        selected_date = datetime.strptime(date, '%Y-%m-%d').date()
        oras = Book_date.query.filter_by(branch='Zapote').with_entities(Book_date.date).all()
        
        date_values = [datetime_obj.date() for datetime_obj, in oras]
        time_values = [datetime_obj.time() for datetime_obj, in oras]
        
        matching_dates = [(date, time) for date, time in zip(date_values, time_values) if date == selected_date]
        #print(matching_dates)

        # Define time slot ranges
        time_slots = [
            (time(8, 0), time(8, 59)),
            (time(9, 0), time(9, 59)),
            (time(10, 0), time(10, 59)),
            (time(11, 0), time(11, 59)),
            (time(12, 0), time(12, 59)),  # 12pm - 1pm
            (time(13, 0), time(13, 59)),  # 1pm - 2pm
            (time(14, 0), time(14, 59)),  # 2pm - 3pm
            (time(15, 0), time(15, 59)),  # 3pm - 4pm
            (time(16, 0), time(16, 59)),  # 4pm - 5pm
            (time(17, 0), time(17, 59)),  # 5pm - 6pm
            (time(18, 0), time(18, 59)),  # 6pm - 7pm
        ]

        available_slots = []  # Initialize the list here

        for start_time, end_time in time_slots:
            slot_available_count = 10  # Initialize slot count for each time slot
            for matching_date, matching_time in matching_dates:
                if start_time <= matching_time <= end_time:
                    #print('decrement')
                    slot_available_count = slot_available_count - 1
            
            available_slots.append((start_time, end_time, slot_available_count))  # Append tuple with slot count

        return render_template('appointment_time_zapote.html', title='Appointment Time', available_slots=available_slots)
    else:
        flash("Before checking the Available Slots you must choose a date", 'warning')

    return render_template('appointment_time_zapote.html', title='Appointment Time')


@app.route('/appointment_time_FCM/', methods=['GET', 'POST']) #Calendar Appointment FCM
def app_time_fcm():
    date1 = request.args.get('date1')

    if date1:
        selected_date = datetime.strptime(date1, '%Y-%m-%d').date()
        oras = Book_date.query.filter_by(branch='FCM').with_entities(Book_date.date).all()
        
        date_values = [datetime_obj.date() for datetime_obj, in oras]
        time_values = [datetime_obj.time() for datetime_obj, in oras]
        
        matching_dates = [(date, time) for date, time in zip(date_values, time_values) if date == selected_date]
        #print(matching_dates)

        # Define time slot ranges
        time_slots = [
            (time(8, 0), time(8, 59)),
            (time(9, 0), time(9, 59)),
            (time(10, 0), time(10, 59)),
            (time(11, 0), time(11, 59)),
            (time(12, 0), time(12, 59)),  # 12pm - 1pm
            (time(13, 0), time(13, 59)),  # 1pm - 2pm
            (time(14, 0), time(14, 59)),  # 2pm - 3pm
            (time(15, 0), time(15, 59)),  # 3pm - 4pm
            (time(16, 0), time(16, 59)),  # 4pm - 5pm
            (time(17, 0), time(17, 59)),  # 5pm - 6pm
            (time(18, 0), time(18, 59)),  # 6pm - 7pm
        ]

        available_slots = []  # Initialize the list here

        for start_time, end_time in time_slots:
            slot_available_count = 10  # Initialize slot count for each time slot
            for matching_date, matching_time in matching_dates:
                if start_time <= matching_time <= end_time:
                    #print('decrement')
                    slot_available_count = slot_available_count - 1
            
            available_slots.append((start_time, end_time, slot_available_count))  # Append tuple with slot count

        return render_template('appointment_time_FCM.html', title='Appointment Time', available_slots=available_slots)

    else:
        flash("Before checking the Available Slots you must choose a date", 'warning')

    return render_template('appointment_time_FCM.html', title='Appointment Time')


def check_available_slots(branch, selected_date):
    date_query = Book_date.query.filter_by(branch=branch).with_entities(Book_date.date).all()

    selected_time = selected_date.time()  # Extract the time component from selected_date
    selected_datetime = datetime.combine(selected_date.date(), selected_time)

    filtered_date_values = [date[0] for date in date_query if date[0].date() == selected_datetime.date()]

    # Define time slot ranges
    time_slots = [
        (time(8, 0), time(8, 59)),
        (time(9, 0), time(9, 59)),
        (time(10, 0), time(10, 59)),
        (time(11, 0), time(11, 59)),
        (time(12, 0), time(12, 59)),  # 12pm - 1pm
        (time(13, 0), time(13, 59)),  # 1pm - 2pm
        (time(14, 0), time(14, 59)),  # 2pm - 3pm
        (time(15, 0), time(15, 59)),  # 3pm - 4pm
        (time(16, 0), time(16, 59)),  # 4pm - 5pm
        (time(17, 0), time(17, 59)),  # 5pm - 6pm
        (time(18, 0), time(18, 59)),  # 6pm - 7pm
    ]

    # Calculate the time slot for the selected_datetime
    selected_time = selected_datetime.time()
    selected_time_slot = None
    for start_time, end_time in time_slots:
        if start_time <= selected_time <= end_time:
            selected_time_slot = (start_time, end_time)
            break

    if selected_time_slot is None:
        # Handle the case where the selected_time is not within any time slot
        return -1  # Invalid time slot

    # Initialize slot_available_count for the selected time slot
    slot_available_count = 10

    # Decrement slot_available_count based on existing appointments at the selected time
    for matching_datetime in filtered_date_values:
        matching_time = matching_datetime.time()
        if selected_time_slot[0] <= matching_time <= selected_time_slot[1]:
            slot_available_count -= 1

    return slot_available_count


@app.route('/book', methods=['GET', 'POST'])
def book_appointment():
    form = AppointmentForm()
    if form.validate_on_submit():
        branch = form.branch.data
        service = form.service.data
        price = parse_price(service)
        service2 = form.service2.data
        service3 = form.service3.data
        date = form.date.data

        # Create a new instance of the Book_date model
        appointment = Book_date(branch=branch, service=service, service2=service2, service3=service3, date=date, user_id=current_user.user_id)

        selected_date = date
        slot_available_count = check_available_slots(branch, selected_date=selected_date)  # Get the available slots count

        print(slot_available_count)

        if slot_available_count == 0:
            flash(f'Cannot book on that date. No slots available. Please check the available slots', 'danger')
            return redirect(url_for('book_appointment'))

        if 'submit_another' in request.form: # Mag add if yung customer accident click submit and return

            # Store the form data in the session
            session['form_data'] = {
                'branch': branch,
                'service': service,
                'price': price,
                'date': date,
                'service2': service2,
                'service3': service3
            }

            # appointment.is_paid = True
            return redirect(url_for('checkout', service=service, price=price, date=date, branch=branch)) # user_id=current_user.user_id ADD USER SA CHECKOUT

        else:
            flash(f'Congratulations for creating an appointment check "MyAppointments" ', 'success')
            # db.session.add(appointment)
            # db.session.commit()
        
            # Redirect to a success page
            return redirect(url_for('book_appointment'))

    return render_template('book.html',title='Booking', form=form)




@app.route('/checkout/')
def checkout():
        # ADD USER SA CHECKOUT
        selected_branch = request.args.get('branch')
        selected_service = request.args.get('service')
        amount = request.args.get('price')
        date_obj = datetime.strptime(request.args.get('date'), '%Y-%m-%d %H:%M:%S') # Convert date string to datetime object
        formatted_date = date_obj.strftime('%B %d, %Y %I:%M %p') 

        return render_template('checkout.html', title='Checkout', selected_service=selected_service, amount=amount, date=formatted_date, selected_branch=selected_branch)

@app.route('/create-checkout-session', methods=['POST'])
def create_checkout_session():
    # Retrieve necessary data for the payment from the request body
    amount = int(request.json['amount']) * 100  # Convert price to cents
    currency = 'PHP'
    description = request.json['description']
    quantity = 1

    # Make a POST request to the PayMongo API to create the Checkout Session
    url = 'https://api.paymongo.com/v1/checkout_sessions'
    payload = {
        "data": {
            "attributes": {
                "line_items": [
                    {
                        "currency": "PHP",
                        "amount": amount,
                        "name": "Product",
                        "quantity": quantity,
                        "description": description
                    }
                ],
                "payment_method_types": ["dob_ubp","dob","card", "gcash","paymaya",],
                "send_email_receipt": True,
                "show_description": True,
                "show_line_items": True,
                "description": description,
                "cancel_url": "http://localhost:5000/book/", # Change it back if using deployment url
                "success_url": "http://localhost:5000/payment-success/" # Change it back if using deployment url 
            }
        }
    }

    headers = {
        "accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": "Basic c2tfdGVzdF9BS05rWEtGTlBubW0yVFl5R2VEQnhkdGg6"  # Not real api
    }


#Basic c2tfdGVzdF9BS05rWEtGTlBubW0yVFl5R2VEQnhkdGg6 not real

    response = requests.post(url, json=payload, headers=headers)
    data = response.json()

    # Retrieve the checkout_url from the response
    checkout_url = data['data']['attributes']['checkout_url']

    # Add the checkout_url to the response data
    data['checkout_url'] = checkout_url

    # Return the response data as JSON
    return jsonify(data)


#sk_test_AKNkXKFNPnmm2TYyGeDBxdth pk_test_KfYzcPSKD92k7YQSwXGPiGva Testing_keys

@app.route('/payment-success/', methods=['POST','GET'])
def payment_success():
    # Retrieve the form data from the session
    form_data = session.get('form_data')

    if form_data:
        # Extract the relevant data from the form_data dictionary
        branch = form_data['branch']
        price = form_data['price']
        date = form_data['date']
        service = form_data['service']
        service2 = form_data['service2']
        service3 = form_data['service3']

        # print(branch)
        # print(price)
        # print(date)
        # print(service)
        
        # ... Rest of your code to process the payment success ...

        # Clear the form data from the session after processing
        session.pop('form_data', None)

        # Create a new instance of the Book_date model
        appointment = Book_date(branch=branch, service=service, service2=service2, service3=service3, date=date, user_id=current_user.user_id)

        # Mark the appointment as paid
        appointment.is_paid = True

        #Commit the appointment to the database
        db.session.add(appointment)
        db.session.commit()

        # Redirect or render a success page
        return render_template('payment_successful_complete.html', title="Success")

    # Redirect to an error page if form_data is not found in the session
    abort (404)


@app.route('/myappointment/')  # myappointments PAGE
@login_required
def history():
    page = request.args.get('page', 1, type=int)
    appointments = Book_date.query.filter_by(user_id=current_user.user_id).order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)
    return render_template('history.html', title='Appointments', appointments=appointments)

@app.route('/logout/')
def logout():
    logout_user()
    return redirect(url_for('index'))


def save_picture(form_picture):
    random_hex = secrets.token_hex(8)
    _, f_ext = os.path.splitext(form_picture.filename)  # jayson.jpg => ('jayson', '.jpg')
    picture_fn = random_hex + f_ext #Using the random_hex as filenames
    picture_path = os.path.join(app.root_path, 'static/profile_pics', picture_fn)

    output_size = (135, 135)
    i = Image.open(form_picture)
    i.thumbnail(output_size)

    i.save(picture_path)

    return picture_fn


@app.route('/account/', methods=['GET', 'POST'])  # ACCOUNT PAGE
@login_required
def account():
    form = UpdateAccountForm()
    if form.validate_on_submit():
        if form.picture.data:
            picture_file = save_picture(form.picture.data)
            current_user.image_file = picture_file
        current_user.fname = form.fname.data
        current_user.lname = form.lname.data
        current_user.gender = form.gender.data
        current_user.age = form.age.data
        current_user.address = form.address.data
        current_user.email = form.email.data
        current_user.contact = form.contact.data

        account_update = AuditTrail(user_id=current_user.user_id, event_type='Account Update', event_description='account successfully updated')
        db.session.add(account_update)
        db.session.commit()

        print("finish updating")
        flash('Your account has been updated!', 'success')
        return redirect(url_for('account'))
    elif request.method == 'GET':
        form.lname.data = current_user.lname
        form.fname.data = current_user.fname
        form.gender.data = current_user.gender
        form.age.data = current_user.age
        form.address.data = current_user.address
        form.email.data = current_user.email
        form.contact.data = current_user.contact
    
    image_file = url_for('static', filename='profile_pics/' + current_user.image_file)
    remove_unused_images()
    return render_template('account.html', title='Account', image_file=image_file, form=form)


@app.route('/update_password', methods=['GET','POST'])
@login_required
def update_password():
    form = UserUpdatePasswordForm()
    if form.validate_on_submit():
        user = User.query.get(current_user.user_id)
        if user and bcrypt.check_password_hash(user.password, form.current_password.data):
            hashed_password = bcrypt.generate_password_hash(form.new_password.data).decode('utf-8')
            user.password = hashed_password

            # print('user.password:', user.password)
            success_update_password = AuditTrail(user_id=user.user_id, event_type='Success Update', event_description='User successfully update their password')
            db.session.add(user)
            db.session.add(success_update_password)
            db.session.commit()
            # print('Database Commit success')


            flash('You Successfully updated your password', 'success')
            return redirect(url_for('account'))
        else:
            flash('Current Password is invalid', 'danger')

    print(form.errors)
    return render_template('user_update_password.html',title='Update Password', form=form)

@app.route('/adminpage/', methods=['GET','POST'])  # Admin Dashboard
@deny_customer_staff
@login_required
def adminpage():
    employee_acc = User.query.filter(User.role.in_(['Admin', 'Super_admin'])).all()
    num_users = len(User.query.filter_by(role='Customer').all())
    num_activate = len(User.query.filter(User.is_active == True, User.role == 'Customer').all())
    num_deactivate = len(User.query.filter(User.is_active == False, User.role == 'Customer').all())
    num_appointment = len(Book_date.query.filter_by(is_done=False).all())
    num_employee = len(User.query.filter(User.role.in_(['Admin', 'Super_admin','Staff'])).all())
    emp_activate = len(User.query.filter(or_(User.role=='Admin', User.role=='Super_admin', User.role=='Staff'), User.is_active==True).all())
    emp_deactivate = len(User.query.filter(or_(User.role=='Admin', User.role=='Super_admin', User.role=='Staff'), User.is_active==False).all())
    num_products = len(Inventory.query.all())
    inventory = Inventory.query.all()  # get all inventory items

    low_stock_products = []
    upcoming_expiry_products = []
    today = datetime.now().date()
    for product in inventory:
        if product.stock < 10:
            low_stock_products.append(product)
        if (product.expiration_date != datetime.today()):
            upcoming_expiry_products.append(product)
    
    logs = AuditTrail.query.order_by(AuditTrail.timestamp.desc()).limit(5).all()

    return render_template('adminpage.html', title='Admin', num_users=num_users, num_activate=num_activate, num_deactivate=num_deactivate, num_appointment=num_appointment,
        num_products=num_products, logs=logs, low_stock_products=low_stock_products, upcoming_expiry_products=upcoming_expiry_products, inventory=inventory,today=today,
        num_employee=num_employee, emp_activate=emp_activate, emp_deactivate=emp_deactivate,employee_acc=employee_acc)

        #num_employee=num_employee, emp_activate=emp_activate, emp_deactivate=emp_deactivate,employee_acc=employee_acc

@app.route('/usermanagement/', methods=['GET'])
@login_required
@deny_customer_staff
def usermanagement():
    form = RegistrationForm()
    page = request.args.get('page', 1, type=int)
    users = User.query.filter_by(role='Customer').order_by(User.date_join.desc()).paginate(page=page, per_page=5)
    return render_template('usermanagement.html', title='Customer', users=users, form=form)

@app.route('/deactivate-user', methods=['POST'])
@deny_customer_staff
@login_required
def deactivate():
    email = request.form['email']
    user = User.query.filter_by(email=email).first()
    if user is not None:
        user.is_active = False
        account_deactivate = AuditTrail(user_id=current_user.user_id, event_type='Account Deactivate', event_description='account successfully deactivated')
        db.session.add(account_deactivate)
        db.session.commit()
        flash('User has been deactivated.', 'success')
    else:
        flash('User not found.', 'warning')
    return redirect(url_for('usermanagement'))


# Delete after 5 years is in the run.py


@app.route('/activate-user', methods=['POST'])
@deny_customer_staff
@login_required
def activate():
    email = request.form['email']
    user = User.query.filter_by(email=email).first()
    if user is not None:
        user.is_active = True
        account_active = AuditTrail(user_id=current_user.user_id, event_type='Account Activated', event_description='account successfully activated')
        db.session.add(account_active)
        db.session.commit()
        flash('User has been activated.', 'success')
    else:
        flash('User not found.', 'warning')
    return redirect(url_for('usermanagement'))

@app.route('/edit-user/<int:user_id>', methods=['GET','POST'])
@deny_customer_staff
@login_required
def edit_user(user_id):
    form = EditUserForm()

    # Retrieve the user from the database based on the provided ID
    user = User.query.get_or_404(user_id) # this method designed to work primary key column

    if form.validate_on_submit():
        user.fname = form.fname.data
        user.lname = form.lname.data
        user.email = form.email.data
        user.contact = form.contact.data

        account_edit_admin = AuditTrail(user_id=current_user.user_id, event_type='Account Edited', event_description='account successfully edited by admin')
        db.session.add(account_edit_admin)
        db.session.commit()

        flash('The account has been updated!', 'success')
        return redirect(url_for('usermanagement'))
    elif request.method == 'GET':
        form.lname.data = user.lname
        form.fname.data = user.fname
        form.email.data = user.email
        form.contact.data = user.contact
    
    # Pass the user object to the edit user template
    print(form.errors)
    return render_template('edit_user.html', form=form,title='Edit User', user_id=user_id)

@app.route('/user_changepass/<int:user_id>', methods=['GET', 'POST'])
@deny_customer_staff
def user_changepass(user_id):
    form = UpdatePasswordForm()

    user = User.query.get(user_id)

    if form.validate_on_submit():
        # Check if the new password is the same as the old password
        if bcrypt.check_password_hash(user.password, form.new_password.data):
            flash('New password cannot be the same as the old password', 'danger')
        else:
            # Updates Customer's password
            new_hashed_password = bcrypt.generate_password_hash(form.new_password.data).decode('utf-8')
            user.password = new_hashed_password

            db.session.commit()
            flash('You successfully updated the employee password', 'success')
            return redirect(url_for('usermanagement', user_id=user_id))
    print(form.errors)
    return render_template('user_changepass.html', form=form,title='User Password', user=user, user_id=user_id)  

@app.route('/create-user', methods=['GET','POST'])
@deny_customer_staff
@login_required
def create_user():
    # if current_user.is_authenticated:
    #     return redirect(url_for('usermanagement'))
    form = RegistrationForm()
    if current_user.is_authenticated:
        if form.validate_on_submit():
            hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            user = User(fname=form.fname.data, lname=form.lname.data, email=form.email.data, password=hashed_password)

            account_created = AuditTrail(user_id=current_user.user_id, event_type='Account Created', event_description='account successfully created by admin')
            db.session.add(account_created)
            db.session.add(user)
            db.session.commit()

            flash('The Account has been created!', 'success')
            print('User created successfully!') # print statement
            return redirect(url_for('usermanagement'))
    print(form.errors)
    return render_template('admin_create_user.html', form=form)


@app.route('/appointment_management/', methods=['GET','POST'])  # Appointment Management
@deny_customer_staff
@login_required
def app_management():
    form = AppointmentForm()
    page = request.args.get('page', 1, type=int)
    appointments = Book_date.query.filter_by(is_done=False).order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)
    user = User.query.all()
    return render_template('appointment_management.html',title='Appointment Management', form=form, appointments=appointments)

@app.route('/book_admin', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def app_management_book():
    form = AppointmentForm()

    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user is None:
            flash('Email does not exist', 'danger')
            return redirect(url_for('app_management'))

        branch = form.branch.data
        service = form.service.data
        service2 = form.service2.data
        service3 = form.service3.data
        date = form.date.data

        appointment = Book_date(user=user, branch=branch, service=service, service2=service2, service3=service3, date=date)

        app_book = AuditTrail(user_id=current_user.user_id, event_type='Appointment Created', event_description='appointment successfully created')
        db.session.add(app_book)
        db.session.add(appointment)
        db.session.commit()

        flash('Appointment Created', 'success')
        return redirect(url_for('app_management'))

    return render_template('appointment_management.html', form=form)


@app.route('/appointment_admin/<int:id>', methods=['POST'])
@deny_customer_staff
def app_management_done(id):
    appointment = Book_date.query.get_or_404(id)
    appointment.is_done = True

    app_done = AuditTrail(user_id=current_user.user_id, event_type='Appointment Done', event_description='appointment is done')
    db.session.add(app_done)
    db.session.commit()

    flash('The Appointment is Done.', 'success')
    return redirect(url_for('app_management'))

@app.route('/download_appointments')
@deny_customer_staff
def download_appointments():
    # Get the appointments from the database
    appointments = Book_date.query.all()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Branch', 'Service', 'Service2' , 'Service3' , 'Date', 'Contact', 'Last Name', 'Done'])

    # Set the width of each column
    for col_index in range(1, 9):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    # Add data to the worksheet
    for appointment in appointments:
        ws.append([appointment.branch, appointment.service, appointment.service2 , appointment.service3 , appointment.date.strftime('%Y-%m-%d %I:%M:%S %p'), appointment.user.contact, appointment.user.lname, appointment.is_done])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    # generate_appointments = AuditTrail(user_id=current_user.user_id, event_type='Records Created', event_description="appointment record successfully downloaded")
    # db.session.add(generate_appointments)
    # db.session.commit()

    return send_file(output, download_name='appointments_record.xlsx', as_attachment=True)


@app.route('/activate_download')
@deny_customer_staff
def download_users():
    # Get the users from the database
    users = User.query.filter_by(role='Customer').all()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['First_name', 'Last_name','Age','Gender','Email', 'Contacts','Address', 'Date_Join', 'Deactivate'])

    # Set the width of each column
    for col_index in range(1, 10):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    #Add a blank row 
    ws.append([])

    # Add data to the worksheet
    for user in users:
        ws.append([user.fname, user.lname, user.age, user.gender, user.email, user.contact, user.address, user.date_join.strftime('%Y-%m-%d %I:%M:%S %p'), user.is_active])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    generate_customers = AuditTrail(user_id=current_user.user_id, event_type='Records Created', event_description="customer's record successfully downloaded")
    db.session.add(generate_customers)
    db.session.commit()

    return send_file(output, download_name='customer_record_activate.xlsx', as_attachment=True)


@app.route('/logging') # audit page
@deny_customer_staff
def audit_trail():
    month = request.args.get('month')
    page = request.args.get('page', 1, type=int)
    if month:
        logs = AuditTrail.query.filter(func.extract('month', AuditTrail.timestamp) == month).order_by(AuditTrail.log_id.desc()).paginate(page=page, per_page=10)
        if not logs.items:
            flash("There's no logs in this month", "info")
    else:
        logs = AuditTrail.query.order_by(AuditTrail.log_id.desc()).paginate(page=page, per_page=10)
    user = User.query.all()
    return render_template('audit_trail.html',title='Audit Trail', user=user, logs=logs)


@app.route('/logs_download')
@deny_customer_staff
@login_required
def download_logs():

    # Get the logs items from the database
    trail = AuditTrail.query.all()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Log Number', 'User', 'Event Type', 'Event Description', 'Time Stamp'])

    # Set the width of each column
    for col_index in range(1, 6):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    #Add a blank row 
    ws.append([])

    # Add data to the worksheet
    for logs in trail:
        ws.append([logs.log_id, logs.user_id, logs.event_type, logs.event_description,logs.timestamp.strftime('%Y-%m-%d %I:%M:%S %p')])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    generate_logs = AuditTrail(user_id=current_user.user_id, event_type='Logs Created', event_description="Logs successfully downloaded")
    db.session.add(generate_logs)
    db.session.commit()

    return send_file(output, download_name='logs_records.xlsx', as_attachment=True, )

@app.route('/customer_search', methods=['GET', 'POST'])
@deny_customer_staff
def customer_search():
    form = RegistrationForm()
    search_term = request.form.get('lname')

    # Fetch the users from the database based on the search term
    if search_term:
        page = request.args.get('page', 1, type=int)
        users = User.query.filter(User.lname.startswith(search_term)).paginate(page=page, per_page=5)


        if not users.items:
            flash(f"No users found with the last name {search_term}", "warning")
            users = User.query.filter_by(role='Customer').order_by(User.date_join.desc()).paginate(page=page, per_page=5)

    else:
        flash("No search term provided. Showing all users.", "warning")
        page = request.args.get('page', 1, type=int)
        users = User.query.filter_by(role='Customer').order_by(User.date_join.desc()).paginate(page=page, per_page=5)

    # Render the template with the users data and the form
    return render_template('usermanagement.html',title='Customer Management', users=users, form=form)


@app.route('/appointment_search', methods=['GET', 'POST'])
@deny_customer_staff
def appointment_search():
    form = AppointmentForm()
    search_term = request.form.get('lname')

    # Fetch the appointments from the database based on the search term
    if search_term:
        page = request.args.get('page', 1, type=int)
        appointments = Book_date.query.join(User).filter(User.lname.startswith(search_term)).paginate(page=page, per_page=5)

        if not appointments.items:
            flash(f"No appointments found for user with last name {search_term}", "warning")
            appointments = Book_date.query.order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)

        # Render the template with the appointments data and the form
        return render_template('appointment_management.html', appointments=appointments, form=form, search_term=search_term)
        
    else:
        flash("No search term provided. Showing all appointments.", "warning")
        page = request.args.get('page', 1, type=int)
        appointments = Book_date.query.order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)

        # Render the template with the appointments data and the form
        return render_template('appointment_management.html',title='Appointment Management', appointments=appointments, form=form)


@app.route('/inventory', methods=['GET','POST'])
@deny_customer_staff
@login_required
def inventory():
    form = InventoryForm()
    page = request.args.get('page', 1, type=int)
    inventory = Inventory.query.order_by(Inventory.inventory_id.desc()).paginate(page=page, per_page=5)
    return render_template('inventory_management.html',title='Inventory', form=form, inventory=inventory)


@app.route('/inventory_create', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def inventory_create():
    form = InventoryForm()
    # print('inventory page')
    if form.validate_on_submit():
        product = form.product.data
        category = form.category.data
        price = form.price.data 
        stock = form.stock.data
        expiration_date = form.expiration_date.data 

        inventory = Inventory(product=product, category=category, price=price, stock=stock, expiration_date=expiration_date)
        db.session.add(inventory)
        db.session.commit()

        flash('Successfully added', 'success')
        return redirect(url_for('inventory'))
    return render_template('inventory_create.html',title='Inventory', form=form)

@app.route('/inventory_edit/<int:inventory_id>', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def inventory_edit(inventory_id):
    form = InventoryForm()

    inventory = Inventory.query.get_or_404(inventory_id)

    if form.validate_on_submit():
        inventory.product = form.product.data
        inventory.category = form.category.data
        inventory.price = form.price.data
        inventory.stock = form.stock.data
        inventory.expiration_date = form.expiration_date.data

        product_edit = AuditTrail(user_id=current_user.user_id, event_type='Product Update', event_description='product update by admin')
        db.session.add(product_edit)
        db.session.commit()

        flash('You successfully update the product', 'success')
        return redirect(url_for('inventory', inventory_id=inventory_id))

    elif request.method == 'GET':
        form.product.data = inventory.product
        form.category.data = inventory.category
        form.price.data = inventory.price
        form.stock.data = inventory.stock
        form.expiration_date.data = inventory.expiration_date
    print(form.errors)
    return render_template('inventory_edit.html', form=form, inventory=inventory)


@app.route('/product_download')
@deny_customer_staff
@login_required
def download_inventory():

    # Get the inventory items from the database
    inventory = Inventory.query.all()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Product', 'Category', 'Price', 'Stock','Expiration Date'])

    # Set the width of each column
    for col_index in range(1, 5):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    #Add a blank row 
    ws.append([])

    # Add data to the worksheet
    for item in inventory:
        ws.append([item.product, item.category, item.price, item.stock, item.expiration_date])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    generate_inventory = AuditTrail(user_id=current_user.user_id, event_type='Inventory Created', event_description="Inventory successfully downloaded")
    db.session.add(generate_inventory)
    db.session.commit()

    return send_file(output, download_name='inventory_records.xlsx', as_attachment=True, )


@app.route('/delete_product', methods=['POST'])
@deny_customer_staff
@login_required
def delete_product():
    product = request.form['product']
    inventory = Inventory.query.filter_by(product=product).first()
    if inventory is not None:
        db.session.delete(inventory)
        product_delete = AuditTrail(user_id=current_user.user_id, event_type='Product Delete', event_description='the product is deleted from the database')
        db.session.add(product_delete)
        db.session.commit()
        flash('Product has been deleted.', 'success')
    else:
        flash('Product not Found', 'warning')
    return redirect(url_for('inventory'))

@app.route('/inventory_search', methods=['GET','POST'])
@deny_customer_staff
@login_required
def inventory_search():
    form = InventoryForm()
    search_term = request.form.get('product')

    # Fetch the products from the database based on the search term
    if search_term:
        page = request.args.get('page', 1, type=int)
        inventory = Inventory.query.filter(Inventory.product.startswith(search_term)).paginate(page=page, per_page=5)

        if not inventory.items:
            flash(f"No products found with the name {search_term}", 'warning')
            inventory = Inventory.query.order_by(Inventory.inventory_id.desc()).paginate(page=page, per_page=5)
    else:
        flash("No search term provided. Showing all products.", "warning")
        page = request.args.get('page', 1, type=int)
        inventory = Inventory.query.order_by(Inventory.id.desc()).paginate(page=page, per_page=5)

    return render_template('inventory_management.html',title='Inventory', inventory=inventory, form=form)

@app.route('/account_admin/', methods=['GET', 'POST'])  # account admin PAGE
@deny_customer_staff
@login_required
def account_admin():
    form = UpdateAccountForm()
    if form.validate_on_submit():
        if form.picture.data:
            picture_file = save_picture(form.picture.data)
            current_user.image_file = picture_file
        current_user.fname = form.fname.data
        current_user.lname = form.lname.data
        current_user.gender = form.gender.data
        current_user.age = form.age.data
        current_user.address = form.address.data
        current_user.email = form.email.data
        current_user.contact = form.contact.data


        account_update = AuditTrail(user_id=current_user.user_id, event_type='Account Update', event_description='account successfully updated')
        db.session.add(account_update)
        db.session.commit()

        flash('Your account has been updated!', 'success')
        return redirect(url_for('account_admin'))
    elif request.method == 'GET':
        form.lname.data = current_user.lname
        form.fname.data = current_user.fname
        form.gender.data = current_user.gender
        form.age.data = current_user.age
        form.address.data = current_user.address
        form.email.data = current_user.email
        form.contact.data = current_user.contact

    image_file = url_for('static', filename='profile_pics/' + current_user.image_file)
    # print(form.errors)
    print(current_user.email)
    return render_template('account_admin.html', title='Account', image_file=image_file, form=form)

@app.route('/update_password_admin', methods=['GET','POST'])
@deny_customer_staff
@login_required
def update_password_admin():
    form = UpdatePasswordForm()
    if form.validate_on_submit():
        user = User.query.get(current_user.user_id)
        if user and bcrypt.check_password_hash(user.password, form.current_password.data):
            hashed_password = bcrypt.generate_password_hash(form.new_password.data).decode('utf-8')
            user.password = hashed_password

            print('user.password:', user.password)
            success_update_password = AuditTrail(user_id=user.user_id, event_type='Success Update', event_description='Admin successfully update their password')
            db.session.add(user)
            db.session.add(success_update_password)
            db.session.commit()
            print('Database Commit success')


            flash('You Successfully updated your password', 'success')
            return redirect(url_for('account_admin'))
        else:
            flash('Current Password is invalid', 'danger')

    # print(form.errors)
    return render_template('admin_update_password.html',title='Update Password', form=form)



@app.route('/employee_management', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def employee_management():
    form = EmployeeForm()
    page = request.args.get('page', 1, type=int)
    users = User.query.filter(User.role.in_(['Super_admin', 'Staff', 'Admin'])).order_by(User.date_join.desc()).paginate(page=page, per_page=5)

    return render_template('employee_management.html',title='Employee Management', users=users, form=form)


@app.route('/employee_create', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def employee_create():
    form = EmployeeForm()

    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        fname = form.fname.data 
        lname = form.lname.data 
        password = hashed_password
        role = form.role.data 
        email = form.email.data

        user = User(fname=fname, lname=lname, password=password, role=role, email=email)

        # employee_created = AuditTrail(user_id=current_user.user_id, event_type='Employee Created', event_description='employee account successfully created by admin')
        # db.session.add(employee_created)
        db.session.add(user)
        db.session.commit()

        employee = Employee(user_id=user.user_id)
        db.session.add(employee)
        db.session.commit()

        flash('Employee Account has been created!', 'success')
        return redirect(url_for('employee_management'))

    return render_template('employee_create.html', form=form)

@app.route('/deactivate-employee', methods=['POST'])
@deny_customer_staff
@login_required
def deactivate_employee():
    email = request.form['email']
    employee = User.query.filter_by(email=email).first()
    if employee is not None:
        employee.is_active = False
        account_deactivate = AuditTrail(user_id=current_user.user_id, event_type='Account Deactivate', event_description='account successfully deactivated')
        db.session.add(account_deactivate)
        db.session.commit()
        flash('User has been deactivated.', 'success')
    else:
        flash('User not found.', 'warning')
    return redirect(url_for('employee_management'))


# CREATE A ACTIVATE ACCOUNT KAGAYA SA USERMANAGEMENT
@app.route('/activate-employee', methods=['POST'])
@deny_customer_staff
@login_required
def activate_employee():
    email = request.form['email']
    employee = User.query.filter_by(email=email).first()
    if employee is not None:
        employee.is_active = True
        account_activate = AuditTrail(user_id=current_user.user_id, event_type='Account Activate', event_description='account successfully activated')
        db.session.add(account_activate)
        db.session.commit()
        flash('User has been Activated.', 'success')
    else:
        flash('User not found.', 'warning')
    return redirect(url_for('employee_management'))

@app.route('/download_employee')
@deny_customer_staff
@login_required
def download_employee():
    # Get the appointments from the database
    roles = ['Super_admin', 'Staff', 'Admin']
    employee = User.query.filter(User.role.in_(roles)).all()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['First Name', 'Last Name', 'Role', 'Gender', 'Age', 'Email', 'Contact', 'Address', 'Activated', 'Date Join'])

    # Set the width of each column
    for col_index in range(1, 11):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    # Add data to the worksheet
    for employee in employee:
        ws.append([employee.fname, employee.lname, employee.role, employee.gender, employee.age, employee.email, employee.contact, employee.address, employee.is_active , employee.date_join.strftime('%Y-%m-%d %I:%M:%S %p')])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    generate_employee = AuditTrail(user_id=current_user.user_id, event_type='Records Created', event_description="employee record successfully downloaded")
    db.session.add(generate_employee)
    db.session.commit()

    return send_file(output, download_name='employee_record.xlsx', as_attachment=True)

@app.route('/employee_search', methods=['GET','POST'])
@deny_customer_staff
@login_required
def employee_search():
    form = EmployeeForm()
    search_term = request.form.get('lname')

    # Fetch the employees from the database based on the search term
    if search_term:
        page = request.args.get('page', 1, type=int)
        users = User.query.filter(User.lname.startswith(search_term)).paginate(page=page, per_page=5)

        if not users.items:
            flash(f"No surname found with the name {search_term}", 'warning')
            users = User.query.filter(User.role.in_(['Super_admin', 'Staff', 'Admin'])).order_by(User.date_join.desc()).paginate(page=page, per_page=5)

    else:
        flash("No search term provided. Showing all employee.", "warning")
        page = request.args.get('page', 1, type=int)
        users = User.query.filter(User.role.in_(['Super_admin', 'Staff', 'Admin'])).order_by(User.date_join.desc()).paginate(page=page, per_page=5)


    return render_template('employee_management.html',title='Employee Management', users=users, form=form)


@app.route('/employee_account/<int:user_id>', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def employee_account(user_id):
    form = UpdateEmployee()

    user = User.query.get_or_404(user_id)
    employee = Employee.query.filter_by(user_id=user_id).first()

    if form.validate_on_submit():
        user.fname = form.fname.data 
        user.lname = form.lname.data 
        user.role = form.role.data 
        user.email = form.email.data
        user.gender = form.gender.data
        user.age = form.age.data 
        user.contact = form.contact.data 
        user.address = form.address.data
        user.barcode_id = form.barcode.data

        if form.file.data:
            file = form.file.data
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            employee.filename = filename

        if employee is not None:
            if form.role.data in ["Admin", "Super_admin"]:
                employee.is_admin = True
            else:
                employee.is_admin = False

        db.session.commit()

        flash(f"Congratulations you successfully updated the employee's account", 'success')
        return redirect(url_for('employee_account',title='Account', user_id=user_id))

    elif request.method == 'GET':
        form.fname.data = user.fname
        form.lname.data = user.lname 
        form.role.data = user.role
        form.email.data = user.email
        form.gender.data = user.gender
        form.age.data = user.age 
        form.contact.data = user.contact
        form.address.data = user.address
        form.barcode.data = user.barcode_id

    return render_template('employee_account.html',title='Account', form=form, user_id=user_id, user=user, employee=employee)


@app.route('/download_document/<path:file_pdf>', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def download_documents(file_pdf):
    directory = app.config['UPLOAD_FOLDER']
    return send_from_directory(directory=directory, path=file_pdf, as_attachment=True)


@app.route('/update_password_employee/<int:user_id>', methods=['GET','POST'])
@deny_customer_staff
@login_required
def update_password_employee(user_id):
    form = UpdatePasswordForm()
    user = User.query.get(user_id)
    if form.validate_on_submit():
        # Check if the new password is the same as the old password
        if bcrypt.check_password_hash(user.password, form.new_password.data):
            flash('New password cannot be the same as the old password', 'danger')
        else:
            # Update the employee's password
            new_hashed_password = bcrypt.generate_password_hash(form.new_password.data).decode('utf-8')
            user.password = new_hashed_password

            db.session.commit()
            flash('You successfully updated the employee password', 'success')
            return redirect(url_for('employee_account', user_id=user_id))

    print(form.errors)
    return render_template('employee_update_password.html', form=form, user=user, user_id=user_id)

def send_reset_email(user):
    token = user.get_reset_token()
    msg = Message('Password Reset Request', sender='nailloungebusiness94@gmail.com', recipients=[user.email])

    with open(os.path.join(os.path.dirname(__file__), 'reset_password.txt')) as file:
            message_body = file.read()


    reset_token = message_body.format(url_for('reset_token', token=token, _external=True))
    msg.body = reset_token
    mail.send(msg)


@app.route("/reset_password", methods=['GET','POST'])
def reset_request():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = RequestResetForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        send_reset_email(user)
        flash('An email has been sent with instruction to reset your password', 'info')
        return redirect(url_for('login'))
    return render_template('reset_request.html', title='Reset Password', form=form)

@app.route("/reset_password/<token>", methods=['GET','POST'])
def reset_token(token):
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    user = User.verify_reset_token(token)
    if user is None:
        flash("That is an invalid or expired request", 'warning')
        return redirect(url_for('login'))
    form = ResetPasswordForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user.password = hashed_password
        db.session.commit()
        flash('Your password has been updated! You are now able to log in', 'success')
        return redirect(url_for('login'))

    return render_template('reset_token.html', title='Reset Password', form=form)


@app.route("/payments/", methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def payments():
    page = request.args.get('page', 1, type=int)
    payments = Book_date.query.filter_by(is_paid=True).order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)
    user = User.query.all()
    return render_template('payments.html',title='Payments', user=user,payments=payments)


@app.route('/payment_search', methods=['GET', 'POST'])
@deny_customer_staff
def payment_search():
    form = AppointmentForm()
    search_term = request.form.get('lname')

    # Fetch the payments from the database based on the search term
    if search_term:
        page = request.args.get('page', 1, type=int)
        payments = Book_date.query.join(User).filter(User.lname.startswith(search_term)).filter(Book_date.is_paid == True).paginate(page=page, per_page=5)

        if not payments.items:
            flash(f"No payments found for user with last name {search_term}", "warning")
            payments = Book_date.query.order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)

        # Render the template with the payments data and the form
        return render_template('payments.html', payments=payments, form=form, search_term=search_term)
        
    else:
        flash("No search term provided. Showing all payments.", "warning")
        page = request.args.get('page', 1, type=int)
        payments = Book_date.query.order_by(Book_date.book_id.desc()).paginate(page=page, per_page=5)

        # Render the template with the paymentss data and the form
        return render_template('payments.html', title='Payments' , payments=payments, form=form)


@app.route('/download_payments', methods=['GET','POST'])
@deny_customer_staff
def download_payments():
    roles = ['Super_admin', 'Staff', 'Admin']
    book_payment = Book_date.query.all()


    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Last Name','Service','Branch', 'Email','Contact', 'Address', 'Paid'])

    # Set the width of each column
    for col_index in range(1, 6):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    # Add data to the worksheet
    for book in book_payment:
        ws.append([book.user.lname,book.service,book.branch, book.user.email, book.user.contact, book.user.address, book.is_paid])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    generate_payments = AuditTrail(user_id=current_user.user_id, event_type='Records Created', event_description="payments record successfully downloaded")
    db.session.add(generate_payments)
    db.session.commit()

    return send_file(output, download_name='payments_record.xlsx', as_attachment=True)

@app.route('/attendance')
@deny_customer_staff
@login_required
def attendance():
    form = AppointmentForm()
    page = int(request.args.get('page', 1))  # Convert the 'page' parameter to an integer
    attendances = Attendance.query.order_by(Attendance.attendance_id.desc()).paginate(page=page, per_page=5)
    return render_template('attendance_employee.html', title='Attendance', attendances=attendances, form=form)


@app.route('/attendance_search', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def attendance_search():
    form = AttendanceForm()  # Create an instance of the AttendanceForm
    
    # Get the search term from the form input
    search_term = request.form.get('lname')

    if search_term:
        # If a search term is provided, perform a filtered search based on the last name
        page = request.args.get('page', 1, type=int)
        
        # Join the Attendance table with the User table and filter based on the last name
        attendance = Attendance.query.join(User).filter(User.lname.ilike(f'{search_term}%')).paginate(page=page, per_page=5)

        if not attendance.items:
            # If no attendance records are found with the given last name, display a warning message
            flash(f"No users found with the last name {search_term}", "warning")
            
            # Retrieve all attendance records, ordered by attendance_id in descending order
            attendance = Attendance.query.order_by(Attendance.attendance_id.desc()).paginate(page=page, per_page=5)

    else:
        # If no search term is provided, display all attendance records
        flash("No search term provided. Showing all Attendance.", "warning")
        page = request.args.get('page', 1, type=int)
        
        # Retrieve all attendance records, ordered by attendance_id in descending order
        attendance = Attendance.query.order_by(Attendance.attendance_id.desc()).paginate(page=page, per_page=5)

    return render_template('attendance_employee.html', attendances=attendance, form=form)


@app.route('/time-in', methods=['POST']) 
def time_in():
    form = TimeInForm()

    if form.validate_on_submit():
        barcode = form.barcode.data

        user = User.query.filter_by(barcode_id=barcode).first()  # Check if the user exists based on the barcode attribute name

        if user is not None:
            attendance = Attendance(user_id=user.user_id, time_in=datetime.now())  # Assuming the user_id attribute in the Attendance model represents the foreign key to the User model
            db.session.add(attendance)
            db.session.commit()

            flash('Employee Time In', 'success')
            return redirect(url_for('attendance'))
        else:
            flash('INVALID BARCODE ID', 'danger')  # Flash an error message if the user does not exist
    print(form.errors)
    return redirect(url_for('attendance'))


@app.route('/time-out', methods=['POST'])
def time_out():
    form = TimeOutForm()

    if form.validate_on_submit():
        barcode = form.barcode.data

        # Retrieve the user based on the barcode ID
        user = User.query.filter_by(barcode_id=barcode).first()

        if user is not None:
            # Find the active attendance record for the user
            attendance = Attendance.query.filter_by(user_id=user.user_id, time_out=None).first()
            if attendance:
                # Record the time-out for the user
                time_out = datetime.now()
                attendance.time_out = time_out

                # Calculate the duration between time in and time out
                duration = time_out - attendance.time_in

                # Calculate the total hours of work
                total_hours = duration.total_seconds() // 3600

                # Calculate the overtime hours
                overtime_hours = max(duration - timedelta(hours=8), timedelta())

                attendance.total_hours = total_hours
                attendance.overtime_hours = overtime_hours.total_seconds() // 3600  # Convert overtime to hours

                # Save the changes to the database
                db.session.commit()
                flash('Employee Time Out', 'success')
            else:
                # No active time-in record found for the user
                flash('No active time-in record found for this user', 'danger')
        else:
            # Invalid barcode ID entered
            flash('Invalid Barcode ID', 'danger')

    # Redirect back to the attendance page
    return redirect(url_for('attendance'))


@app.route('/download_attendance') 
@login_required
@deny_customer_staff
def download_attendance():
    # Get the attendance from the database
    attendance = Attendance.query.all()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['First Name', 'Last Name', 'Role', 'Time In', 'Time Out'])

    # Set the width of each column
    for col_index in range(1, 11):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    # Add data to the worksheet
    for attendance in attendance:
        ws.append([attendance.user.fname, attendance.user.lname, attendance.user.role, attendance.time_in.strftime('%Y-%m-%d %I:%M:%S %p') if attendance.time_in else "", attendance.time_out.strftime('%Y-%m-%d %I:%M:%S %p') if attendance.time_out else ""])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    # generate_employee = AuditTrail(user_id=current_user.user_id, event_type='Records Created', event_description="employee record successfully downloaded")
    # db.session.add(generate_employee)
    # db.session.commit()

    return send_file(output, download_name='attendance_record.xlsx', as_attachment=True)

##CALCULATE THE PERSONS OVERALL TOTAL HOURS OF WORK AND OVERTIME##
def calculate_overall_total_hours(user_id, month):
    attendances = Attendance.query.filter(
        Attendance.user_id == user_id,
        extract('month', Attendance.time_in) == month
    ).all()

    total_hours = 0
    total_overtime = 0

    for attendance in attendances:
        if attendance.time_out:
            duration = attendance.time_out - attendance.time_in
            total_hours += duration.total_seconds() / 3600
            overtime_hours = max(duration - timedelta(hours=8), timedelta())
            total_overtime += overtime_hours.total_seconds() / 3600

    return total_hours, total_overtime


def calculate_gross_and_net_pay(total_hours, total_overtime, payrate, tax):
    overtime_rate = 1.5
    gross_pay = (total_hours * payrate) + (total_overtime * overtime_rate)
    tax_amount = gross_pay * (tax / 100)
    net_pay = gross_pay - tax_amount

    return gross_pay, net_pay

@app.route('/payroll/<int:user_id>', methods=['GET', 'POST'])
@deny_customer_staff
@login_required
def payroll(user_id):
    form = PayrollForm()

    user = User.query.get_or_404(user_id)
    employee = Employee.query.filter_by(user_id=user_id).first()
    full_name = f"{user.fname} {user.lname}"

    # Get the current month
    current_month = datetime.now().month

    # Calculate the overall total hours and overtime hours
    total_hours, total_overtime = calculate_overall_total_hours(user_id, current_month)

    if form.validate_on_submit():
        # Update the employee's data with the form inputs
        employee.full_name = full_name
        employee.payrate = form.payrate.data
        employee.tax = form.tax.data
        employee.month = current_month

        # Calculate gross pay and net pay using the calculated total hours and overtime
        gross_pay, net_pay = calculate_gross_and_net_pay(
            total_hours, total_overtime,
            form.payrate.data, form.tax.data
        )

        # Create  the payroll object associated with the employee
        payroll = Payroll.query.filter_by(user_id=user_id).first()

        payroll = Payroll(payrate=form.payrate.data,gross_pay=gross_pay, tax=form.tax.data,net_pay=net_pay,month=form.month.data,user_id=user_id, generated=datetime.now())

        try:
            db.session.add(payroll)
            db.session.commit()
            print("Database updated")
            flash('Payroll information created successfully', 'success')
        except Exception as e:
            db.session.rollback()
            print("Database update failed:", str(e))
            flash('Error creating payroll information', 'danger')

        return redirect(url_for('payroll', user_id=user_id))

    elif request.method == 'GET':
        form.full_name.data = full_name
        form.total_hours.data = total_hours
        form.total_overtime.data = total_overtime
        # Assign other form fields with existing data from the database
    print(form.errors)
    return render_template('payroll.html',title='Payroll', user_id=user_id, form=form, total_hours=total_hours, total_overtime=total_overtime)


@app.route('/download_payroll/<int:user_id>', methods=["GET","POST"])
@deny_customer_staff
@login_required
def download_payroll(user_id):
    # Get the specific payroll from the database
    payroll = Payroll.query.filter_by(user_id=user_id).all()

    current_date_time = datetime.now()

    # Create a new workbook and a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['First Name', 'Last Name', 'Month','PayRate', 'Tax', 'Gross Pay', 'Net Pay', "Generated"])

    # Set the width of each column
    for col_index in range(1, 11):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = 30  # 15 characters

    # Add data to the worksheet
    for payroll in payroll:
        generated_formatted = payroll.generated.strftime("%Y-%m-%d %H:%M:%S")
        ws.append([payroll.user.fname, payroll.user.lname, payroll.month, payroll.payrate, payroll.tax, payroll.gross_pay, payroll.net_pay, generated_formatted])

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)

    # Create a response object with the workbook as an attachment
    output.seek(0)

    # generate_employee = AuditTrail(user_id=current_user.user_id, event_type='Records Created', event_description="employee record successfully downloaded")
    # db.session.add(generate_employee)
    # db.session.commit()

    return send_file(output, download_name='user_payroll_record.xlsx', as_attachment=True)


# with app.app_context():
#     new_employee = User(
#         fname='root',
#         lname='root',
#         password=bcrypt.generate_password_hash('Jayson$$$21').decode('utf-8'),  # hash the password and store it,
#         gender='Male',
#         age=35,
#         email='r00t@Adm1n.com',
#         role = 'Super_admin',
#         contact='09568866863',
#         address='123 Main St, Anytown, USA',
#         date_join=datetime.now(),
#         is_active=True,
#         image_file='default.jpg'
#     )
#     db.session.add(new_employee)
#     db.session.commit()


# with app.app_context():
#     new_row = Employee(
#         employee_id = 1,
#         user_id = 1,
#         is_admin =1,
#         filename= None
#         )
#     db.session.add(new_row)
#     db.session.commit()